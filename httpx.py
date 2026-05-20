# 实现:  5.批量url探活+链接提取
import os.path

import openpyxl

标题 = "\033[33m" + r"""
.__       __     __                       ________  
|  |__  _/  |_ _/  |_ ______  ___  ___    \\_____  \\ 
|  |  \ \   __\   __\____ \ \  \/  /     /  ____/ 
|   Y  \ |  |   |  |  |  |_> > >    <     /       \ 
|___|  / |__|   |__|  |   __/ /__/\_ \ /\ \_______ \
     \/               |__|          \/ \/         \/
"""+ "\033[0m"+"\033[31m"+r"""
[*]version:3.0.0
"""

print(标题)

import requests
import random
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

ua=[
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36 Edg/146.0.0.0',
    'Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.84 Safari/537.36'
]


#这里是用来读取文件的
class file_read():

    def file_url(self):
        file_name="url.txt"
        with open(file_name,'r',encoding='utf-8') as f:
            urls=f.read().splitlines()
        return sorted(list(urls))    #这个用来返回文件中的url

    def file_classified(self):  #来分类这些url
        url_s=self.file_url()
        # print(url_s)
        """
        一般分为 ip/纯域名、加http的、加https的、域名
        """
        http=[]    #放http
        https=[]    #放https
        ip=[]    #放ip/域名
        for url in url_s:
            if url.startswith('http://'):
                http.append(url)
                https.append(url.replace('http://', 'https://'))
                # ip.append(url.replace('http://', ''))
            elif url.startswith('https://'):
                https.append(url)
                http.append(url.replace('https://', 'http://'))
                # ip.append(url.replace('https://', ''))
            else:
                # ip.append(url)    #不要纯域名的
                https.append('https://'+url)
                http.append('http://'+url)
        total=http+https+ip    #用来放总的
        total_url=sorted(list(set(total)))
        return total_url



#请求+页面链接提取
class request_url():
    def __init__(self):
        self.url = None
        self.content = None

    def req_url(self, url, retries=2):
        self.url = url
        protocol = url.split('://')[0] if '://' in url else '-'
        for attempt in range(retries + 1):
            try:
                response = requests.get(url, headers={'User-Agent': random.choice(ua)}, timeout=5)
                response.encoding = 'utf-8'
                self.content = response.text
                try:
                    title_match = re.findall(r'<title>(.*?)</title>', response.text, re.IGNORECASE)
                    title = title_match[0] if title_match else '-'
                except Exception:
                    title = '-'
                server = str(response.headers.get('Server', '-'))
                return [url, response.url, self.extract_host(url), response.status_code, len(response.text), title, server, protocol]
            except Exception:
                if attempt < retries:
                    continue
                self.content = None
                return [url, url, self.extract_host(url), 0, 0, '-', '-', protocol]

    def link_url(self):
        # 如果内容为空或None，直接返回空结果
        if not self.content:
            return self.domain_url(), []

        # HTML属性中的链接
        re_rules=[
            r'href=([\'"])(.*?)\1',           # href链接
            r'src=([\'"])(.*?)\1',            # src资源
            r'action=([\'"])(.*?)\1',         # 表单提交地址
            r'data-src=([\'"])(.*?)\1',       # 懒加载图片
            r'data-href=([\'"])(.*?)\1',      # 自定义跳转
            r'data-url=([\'"])(.*?)\1',       # 自定义URL
            r'srcset=([\'"])(.*?)\1',         # 响应式图片
            r'background=([\'"])(.*?)\1',     # 背景图
            r'poster=([\'"])(.*?)\1',         # 视频封面
            r'content=([\'"])\d+;url=(.*?)\1', # meta跳转
        ]
        url=[]
        for rule in re_rules:
            url+=re.findall(rule, self.content, re.IGNORECASE)

        # CSS中的url()
        css_urls=re.findall(r'url\(([\'"]?)(.*?)\1\)', self.content, re.IGNORECASE)
        url+=[u for u in css_urls if u[1]]

        # JavaScript中的链接
        js_rules=[
            r'(?:window\.open|location\.href|location\.assign)\s*\(\s*([\'"])(.*?)\1',
            r'(?:src|href)\s*[:=]\s*([\'"])(.*?)\1',
        ]
        for rule in js_rules:
            url+=re.findall(rule, self.content, re.IGNORECASE)
        url_list=[]
        for item in url:
            link = item[1].strip()
            if not link or link.startswith('#') or link.startswith('javascript:'):
                continue

            # 过滤和规范化
            link = link.split('#')[0].split('?')[0] if '#' in link or '?' in link else link
            link = re.sub(r'[\'">\s]+$', '', link)  # 清理尾部杂质

            if link.startswith('http://'):
                houzui=link[7:].replace('///','/').replace('//','/')
                url_list.append('http://'+houzui)
            elif link.startswith('https://'):
                houzui=link[8:].replace('///','/').replace('//','/')
                url_list.append('https://'+houzui)
            elif link.startswith('//'):
                url_list.append('https://'+link[2:])
            elif link.startswith('/'):
                houzui=link.replace('///','/').replace('//','/')
                url_list.append(self.domain_url()+houzui)

        return self.domain_url(),sorted(list(set(url_list)))

        # 返回域名
    def domain_url(self):
        if self.url.startswith('http:'):
            domain_u=re.findall(r'http://[^/]+',self.url)
        elif self.url.startswith('https:'):
            domain_u=re.findall(r'https://[^/]+',self.url)
        else:
            domain_u=['null']
        return domain_u[0]

    # 提取host（域名:端口）
    def extract_host(self, url):
        try:
            if url.startswith('http://'):
                host_part = url[7:].split('/')[0].split('?')[0]
            elif url.startswith('https://'):
                host_part = url[8:].split('/')[0].split('?')[0]
            else:
                host_part = url.split('/')[0].split('?')[0]
            return host_part
        except:
            return 'null'

    def display(self,response):
        #['请求url', '响应url', 'host', 响应码, 响应长度, 'title', 'Server', 'protocol']
        if str(response[3]).startswith('2'):
            print(f'\033[32m[{response[3]}] -- [{response[4]}] {response[1]}  [title]:{response[5]}   [server]:{response[6]}   [protocol]:{response[7]}\033[0m')
        elif str(response[3]).startswith('3'):
            print(f'\033[34m[{response[3]}] -- [{response[4]}] {response[1]}  [title]:{response[5]}   [server]:{response[6]}   [protocol]:{response[7]}\033[0m')
        elif str(response[3]).startswith('4'):
            print(f'\033[31m[{response[3]}] -- [{response[4]}] {response[1]}  [title]:{response[5]}   [server]:{response[6]}   [protocol]:{response[7]}\033[0m')
        elif str(response[3]).startswith('5'):
            print(f'\033[33m[{response[3]}] -- [{response[4]}] {response[1]}  [title]:{response[5]}   [server]:{response[6]}   [protocol]:{response[7]}\033[0m')
        elif str(response[3]).startswith('0'):
            print(f'\033[35m[{response[3]}] -- [{response[4]}] {response[1]}  [title]:{response[5]}   [server]:{response[6]}   [protocol]:{response[7]}\033[0m')
        else:
            print(f'[{response[3]}] -- [{response[4]}] {response[1]}  [title]:{response[5]}   [server]:{response[6]}   [protocol]:{response[7]}')


class save():

    def save (self,url_list):
        if not os.path.exists('result'):
            os.mkdir('result')
        if not os.path.exists('result/url.xlsx'):
            wb=openpyxl.Workbook()
            wb.save('result/url.xlsx')
        if os.path.getsize('result/url.xlsx')<=10240:
            wb=openpyxl.load_workbook('result/url.xlsx')
            ws=wb.active
            ws.title='url'
            if ws['A1'].value==None:
                ws['A1']='request_url'
                ws['B1']='response_url'
                ws['C1']='host'
                ws['D1']='status_code'
                ws['E1']='length'
                ws['F1']='title'
                ws['G1']='server'
                ws['H1']='protocol'
                wb.save('result/url.xlsx')
        wb = openpyxl.load_workbook('result/url.xlsx')
        ws = wb.active
        for url in url_list:
            ws.append(url)
        wb.save('result/url.xlsx')

    def save_source(self,source_url_list):
        if not os.path.exists('result'):
            os.mkdir('result')
        for domain_url,source_url_list in source_url_list.items():
            with open(f'result/{domain_url.replace('http://','').replace('https://','')}.txt', 'a', encoding='utf-8') as f:
                for url in source_url_list:
                    f.write(url+'\n')

url=file_read().file_classified()  #获取全部url（处理过了）
url_list=[]     #用来存储原来的url的响应

source_url_list={}  #源码中爬取到的链接

# 用于跟踪每个域名的协议信息
protocol_tracker = {}
processed_hosts = {}  # 避免重复处理同一域名

# 线程锁，保护共享数据和print输出
lock = threading.Lock()

# 线程数，默认10，可通过环境变量 THREADS 调整
THREADS = int(os.environ.get('THREADS', 10))

# 统计计数器
stats = {'total': len(url), 'done': 0, 'success': 0, 'timeout': 0, 'filtered': 0}

def process_url(u):
    """处理单个url：请求 + 链接提取"""
    req = request_url()
    response1 = req.req_url(u)

    # 统计请求失败（状态码0=超时/连接失败）
    if response1[3] == 0:
        return {'status': 'timeout', 'url': u, 'host': req.extract_host(u)}

    # 提取链接（每个req实例独立，线程安全）
    domain_url, source_url = req.link_url()

    return {
        'status': 'ok',
        'response': response1,
        'host': response1[2],
        'protocol': response1[7],
        'domain_url': domain_url,
        'source_url': source_url,
    }

print(f'\033[36m[*] 线程数: {THREADS}, 总URL数: {stats["total"]}\033[0m')

with ThreadPoolExecutor(max_workers=THREADS) as executor:
    futures = {executor.submit(process_url, u): u for u in url}
    for future in as_completed(futures):
        result = future.result()
        with lock:
            stats['done'] += 1
            # 每处理20个URL打印一次进度
            if stats['done'] % 20 == 0 or stats['done'] == stats['total']:
                print(f'\033[36m[*] 进度: {stats["done"]}/{stats["total"]}'
                      f'  成功: {stats["success"]}  超时: {stats["timeout"]}'
                      f'  唯一host: {len(protocol_tracker)}\033[0m')

        if result['status'] == 'timeout':
            with lock:
                stats['timeout'] += 1
            continue

        with lock:
            stats['success'] += 1
        host = result['host']
        protocol = result['protocol']
        response1 = result['response']

        with lock:
            # 跟踪协议信息
            if host not in protocol_tracker:
                protocol_tracker[host] = {'http': False, 'https': False, 'responses': []}

            if protocol == 'http':
                protocol_tracker[host]['http'] = True
            elif protocol == 'https':
                protocol_tracker[host]['https'] = True

            protocol_tracker[host]['responses'].append(response1)

            # 只在首次处理该域名时提取链接，避免重复
            if host not in processed_hosts:
                req = request_url()
                req.display(response1)
                source_url_list[result['domain_url']] = sorted(list(set(result['source_url'])))
                processed_hosts[host] = True

# 处理协议信息并生成最终结果
final_results = []
for host, info in protocol_tracker.items():
    # 检查是否有协议跳转情况
    http_to_https = False
    https_to_http = False

    # 分析跳转情况
    for response in info['responses']:
        request_protocol = response[0].split('://')[0] if '://' in response[0] else ''
        response_protocol = response[1].split('://')[0] if '://' in response[1] else ''

        if request_protocol == 'http' and response_protocol == 'https':
            http_to_https = True
        elif request_protocol == 'https' and response_protocol == 'http':
            https_to_http = True

    # 根据跳转情况决定保留哪些结果
    if http_to_https and not https_to_http:
        # http跳转到https，只保留https结果
        for response in info['responses']:
            if response[1].startswith('https://'):
                response[7] = 'https'
                final_results.append(response)
    elif https_to_http and not http_to_https:
        # https跳转到http，只保留http结果
        for response in info['responses']:
            if response[1].startswith('http://'):
                response[7] = 'http'
                final_results.append(response)
    else:
        # 没有跳转或双向跳转，按原逻辑处理
        if info['http'] and info['https']:
            protocol_str = 'http / https'
        elif info['http']:
            protocol_str = 'http'
        elif info['https']:
            protocol_str = 'https'
        else:
            protocol_str = '-'

        for response in info['responses']:
            response[7] = protocol_str
            final_results.append(response)

url_list = final_results

# 最终统计
print(f'\n\033[36m{"="*50}')
print(f'[*] 统计结果:')
print(f'    输入URL数: {stats["total"]}')
print(f'    请求成功: {stats["success"]}')
print(f'    请求超时/失败: {stats["timeout"]}')
print(f'    唯一host数: {len(protocol_tracker)}')
print(f'    最终输出行数: {len(url_list)}')
print(f'    提取链接域名数: {len(source_url_list)}')
print(f'{"="*50}\033[0m\n')

#保存基本的探活文件
save_file=save()
save_file.save(url_list)

#保存额外提取的源代码链接
save_file.save_source(source_url_list)
